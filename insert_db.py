import pymysql
from openpyxl import load_workbook

# DB 연결
con = pymysql.connect(host='db-hotmovie.cv0xzoma7waa.ap-northeast-2.rds.amazonaws.com', user='admin', password='hotmovie',db='k5smovie', charset='utf8') # 한글처리 (charset = 'utf8')

# 커서 만들기
cur = con.cursor()

# 영화 정보 엑셀 파일 읽기
wb = load_workbook("k5smovie.xlsx", data_only=True)
ws = wb.active
real_list = []
for x in range(2, ws.max_row+1):
	list = []
	for y in range(1, 12):
		if ws.cell(row=x, column=y).value == None:
			break
		else:
			if y == 8:
				tmp_str = ws.cell(row=x, column=y).value
				tmp_str = tmp_str.replace("_x000D_\xa0"," ")
				tmp_str = tmp_str.replace("\xa0"," ")
				list.append(tmp_str)
			else:
				list.append(ws.cell(row=x, column=y).value)
	if len(list) != 0:
		list.append(ws.cell(row=x, column=y).value)
		real_list.append(list)

# # sql 파일 참고
# sql = ""
# cur.execute(sql)

# 영화 정보 DB에 삽입
sql = "insert into mv_table values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
# s3 링크 https://hotmovie.s3.ap-northeast-2.amazonaws.com/imgs/190695.png
path = "https://hotmovie.s3.ap-northeast-2.amazonaws.com/imgs/"

for i in range(len(real_list)):
	path2 = path + real_list[i][0] + ".png"
	vals = (real_list[i][0], real_list[i][1], real_list[i][2], real_list[i][3], real_list[i][4], real_list[i][5], real_list[i][6], real_list[i][7], real_list[i][8],  real_list[i][9],  real_list[i][10], path2)
	cur.execute(sql, vals)

rows = cur.fetchall()
print("sql문 결과")
print(rows)     # 전체 rows

# 커밋을 안해주면 DB에 반영되지 않음
# con.commit()

# 연결 종료
con.close()